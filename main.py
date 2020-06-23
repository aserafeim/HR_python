# from deformation_version2 import strain_calc, calc_deformation_dislocation
import openpyxl
from deformation_pass import deformation_and_drx, strain_calc
import static_recrystallisation
import inputdata
import material_constants as mat_const
import math
import time
from exporter import export_to_excel, export_pass, export_by_pass
from plotting import plotting_step
from plotting2 import plot_pass
import numpy
import os

start = time.time()
location = r'D:\HotRolling-env\Check_HR\Test_srx.xlsx'
wb = openpyxl.load_workbook(location)
sheet_curr = wb['Static_RX']
max_col = sheet_curr.max_column
max_row = sheet_curr.max_row
main_dict = {}

'''reading excel file'''
for i in range(3, max_row + 1):
    temp_dict = {}
    temp_dict['time_increment'] = sheet_curr.cell(i, 2).value
    temp_dict['initial_strain'] = sheet_curr.cell(i, 3).value
    temp_dict['final_strain'] = sheet_curr.cell(i, 4).value
    temp_dict['initial_strain_rate'] = sheet_curr.cell(i, 5).value
    temp_dict['final_strain_rate'] = sheet_curr.cell(i, 6).value
    temp_dict['initial_temp'] = sheet_curr.cell(i, 7).value + 273.15
    temp_dict['final_temp'] = sheet_curr.cell(i, 8).value + 273.15
    temp_dict['rho_0'] = sheet_curr.cell(i, 9).value if (isinstance(sheet_curr.cell(i, 9).value, int) or
                                                         isinstance(sheet_curr.cell(i, 9).value, float)) else 0.0
    temp_dict['pass_interval'] = sheet_curr.cell(i, 10).value
    temp_dict['temp_rate_of_change'] = (temp_dict['final_temp'] - temp_dict['initial_temp'])/temp_dict['pass_interval']
    ''
    if temp_dict['initial_strain_rate'] != 0.0:
        temp_dict['reference'] = 'deformation'
    else:
        temp_dict['reference'] = 'interpass'
    main_dict[i - 2] = temp_dict

'start of processing of each pass'
counter = 0
'storage dictionary for storing values of the working dictionary which will be ' \
'updated at the end to the variable pass dictionary'

storage_dict = {'strain': [],
                'temperature': [],
                'time': [],
                'rho_def': [],
                'global_time': [],
                'r_rx_prev_critical': [],
                'r_rx': [],
                'r_g_curr': [],
                'n_rx': [],
                'd_nrx_dt': [],
                'x_curr': [],
                'stress': [],
                'driving_force': [],
                'x_c_curr': [],
                'd_mean': [],
                'drg_dt': [],
                'n_d': [],
                'rho_m': [],
                'rho_critical': [],
                'd_nrx': [],
                'rho_rxed': [],
                'r_critical': [],
                'D_def': []
}
'''running of loops for passes'''
n_d = 0.0

for i in range(1, max_row - 1):

    ''' initialisation of variables '''
    variable_pass = main_dict[i]

    time_step = variable_pass['time_increment']
    strain_initial = variable_pass['initial_strain']
    pass_interval = variable_pass['pass_interval']
    time_curr = time_step
    strain_rate_initial = variable_pass['initial_strain_rate']
    strain_rate_temp = strain_rate_initial

    '''1st step and if deformation'''
    if i == 1:
        storage_dict['rho_def'].append(variable_pass['rho_0'])
        storage_dict['n_rx'].append(0.0)
        storage_dict['r_g_curr'].append(0.0)
        storage_dict['x_curr'].append(0.0)
        storage_dict['x_c_curr'].append(0.0)
        storage_dict['r_rx'].append(0.0)
        storage_dict['d_mean'].append(inputdata.d0)
        storage_dict['D_def'].append(inputdata.d0)
        'added to make the array of the same length'
        storage_dict['time'].append(0.0)
        storage_dict['rho_m'].append(1.02e11)
        storage_dict['stress'].append(0.0)
        storage_dict['rho_critical'].append(1e11)

    ''' last_time used for real time calculation and global time calculation'''
    last_pass_interval = main_dict[i - 1]['global_time'][-1] if i > 1 else 0.0

    storage_dict['temperature'].append(variable_pass['initial_temp'])
    storage_dict['strain'].append(strain_initial)
    print(" ----------------- pass ", i, " ----------------------")
    '''checks if strain rate is 0 or not, i.e deformation or static RX'''
    if strain_rate_initial > 0:
        'sub-routine counter variable'
        j = 0
        d_m_const = storage_dict['d_mean'][0]
        print(storage_dict)
        while time_curr <= pass_interval:
            strain_jplus1 = strain_calc(storage_dict['strain'][j], time_step, strain_rate_initial)
            storage_dict['strain'].append(strain_jplus1)
            storage_dict['temperature'].append(storage_dict['temperature'][j] +
                                               variable_pass['temp_rate_of_change'] * time_step)

            'Call to the deformation module'
            deformation_dict = deformation_and_drx(storage_dict['temperature'][j],
                                                   strain_rate_initial, time_step, storage_dict['rho_def'][j], storage_dict['x_curr'][j],
                                                   storage_dict['r_g_curr'][j],
                                                   storage_dict['n_rx'][j], storage_dict['x_c_curr'][j], (strain_jplus1-strain_initial), d_m_const)

            'Adding values to the lists of temp, strain and dislocation_density'
            storage_dict['time'].append(time_curr)

            'appending values to the lists'
            storage_dict['r_critical'].append(deformation_dict['r_xc'])
            storage_dict['rho_def'].append(deformation_dict['rho_def'])
            storage_dict['x_curr'].append(deformation_dict['x_curr'])
            storage_dict['rho_m'].append(deformation_dict['rho_m'])
            storage_dict['r_rx'].append(deformation_dict['r_rx'])
            storage_dict['r_g_curr'].append(deformation_dict['r_g_curr'])
            storage_dict['d_nrx_dt'].append(deformation_dict['d_nrx_dt'])
            storage_dict['n_rx'].append(deformation_dict['n_rx'])
            storage_dict['d_mean'].append(deformation_dict['d_mean'])
            storage_dict['x_c_curr'].append(deformation_dict['x_c_curr'])
            storage_dict['drg_dt'].append(deformation_dict['drg_dt'])
            storage_dict['rho_rxed'].append(deformation_dict['rho_rxed'])
            storage_dict['n_d'].append(deformation_dict['n_d'])
            storage_dict['D_def'].append(deformation_dict['D_def'])
            storage_dict['d_nrx'].append((deformation_dict['d_nrx']))
            storage_dict['rho_critical'].append(deformation_dict['rho_critical'])
            
            if deformation_dict['x_curr'] > 0.99:
                # plot(storage_dict['time'], storage_dict['x_curr'], 'Time', 'drx_xcurr', True, False,
                #               variable_pass)
                print("rx fraction greater than 1 ", deformation_dict)
                break
            if deformation_dict['rho_def'] <= 0:
                print(j)
                print(deformation_dict)
                export_pass(storage_dict, '1st_pass_err')
            'calculation of stress'
            storage_dict['stress'].append(mat_const.calculate_stress(storage_dict['temperature'][j], strain_rate_temp, deformation_dict['rho_m']))
            'Updating variables for next iteration of the sub-routine'
            time_curr += time_step
            j += 1
    else:
        j = 0
        'Static recrystallisation module'
        print(storage_dict)
        d_m_const = storage_dict['d_mean'][0]
        while time_curr <= pass_interval:

            'calculation of strain and temperature'
            strain_jplus1 = strain_calc(storage_dict['strain'][j], time_step, strain_rate_initial)
            storage_dict['strain'].append(strain_jplus1)
            storage_dict['temperature'].append(storage_dict['temperature'][j] +
                                                   variable_pass['temp_rate_of_change'] * time_step)

            'call to static module'
            temp_srx_dict = static_recrystallisation.static_rx(storage_dict['rho_m'][j], storage_dict['rho_def'][j],
                                                                storage_dict['n_rx'][j], storage_dict['r_g_curr'][j],
                                                                storage_dict['d_mean'][j], storage_dict['temperature'][j], time_step,
                                                                storage_dict['r_rx'][j], d_m_const, storage_dict['x_curr'][j], storage_dict['D_def'][j], strain_jplus1)

            if temp_srx_dict['rho_m'] <= 0:
                print(j)
                print(temp_srx_dict)
                # export_pass(temp_srx_dict, '1st_pass_err')
            'Stress calculation'
            storage_dict['stress'].append(mat_const.calculate_stress(storage_dict['temperature'][j], strain_rate_temp,
                                                                         temp_srx_dict['rho_m']))
            'adding values to the list, later used for plotting the graphs'
            storage_dict['time'].append(time_curr)
            storage_dict['d_mean'].append(temp_srx_dict['d_mean'])
            storage_dict['rho_m'].append(temp_srx_dict['rho_m'])
            storage_dict['n_rx'].append(temp_srx_dict['n_rx'])
            storage_dict['x_curr'].append(temp_srx_dict['x_curr'])
            storage_dict['d_nrx_dt'].append(temp_srx_dict['d_nrx_dt'])
            storage_dict['d_nrx'].append(temp_srx_dict['d_nrx'])
            storage_dict['r_g_curr'].append(temp_srx_dict['r_g_curr'])
            storage_dict['driving_force'].append(temp_srx_dict['driving_force'])
            storage_dict['r_rx'].append(temp_srx_dict['r_rx'])
            storage_dict['r_critical'].append(temp_srx_dict['r_critical'])
            storage_dict['r_rx_prev_critical'].append(temp_srx_dict['r_rx_prev_critical'])
            storage_dict['rho_def'].append(temp_srx_dict['rho_def'])
            storage_dict['D_def'].append(temp_srx_dict['D_def'])
            storage_dict['drg_dt'].append(temp_srx_dict['drg_dt'])

            'updating values after every iteration'
            time_curr += time_step
            j += 1


        storage_dict['n_d'] = [n_d] * len(storage_dict['time'])

    'keys and values of storage dict for each iteration is added to the variable pass, which contains the '
    for key in storage_dict:
        variable_pass[key] = storage_dict[key].copy()

    # if i==3:
    #     export_pass(storage_dict, '3rd_pass_Def')
    for key, val in storage_dict.items():
        if type(val) == list and len(val) > 0:
            val_sc = "{:e}".format(val[-1])
            print(key, ": ", val_sc)

    # workbook = export_by_pass(storage_dict, workbook, i)

    for value in storage_dict.values():
        del value[:]

    for x in variable_pass['time']:
        variable_pass['global_time'].append(last_pass_interval + x)

    for key, val in variable_pass.items():
        if key == 'stress' or key == 'global_time' or key == 'strain':
            print(key, " : ", len(variable_pass[key]))

    # if i == 3:
    #     export_pass(variable_pass, "2nd_pass_interpass")

    if i < max_row - 2:
        storage_dict['d_mean'].append(variable_pass['d_mean'][-1])
        storage_dict['stress'].append(variable_pass['stress'][-1])
        storage_dict['time'].append(0.0)

        'special conditions if deformation -> interpass and is the first step'
        if variable_pass['reference'] == 'deformation' and main_dict[i+1]['reference'] == 'interpass':
            storage_dict['rho_m'].append(variable_pass['rho_m'][-1])
            storage_dict['rho_def'].append(variable_pass['rho_def'][-1])
            storage_dict['r_rx'].append(0.5 * variable_pass['d_mean'][-1])
            storage_dict['n_rx'].append(variable_pass['n_rx'][-1])
            storage_dict['D_def'].append(variable_pass['D_def'][-1])
            storage_dict['x_curr'].append(0.0)
            storage_dict['r_g_curr'].append(variable_pass['r_g_curr'][-1])
            n_d = variable_pass['n_d'][-1]

        'interpass to interpass'
        if variable_pass['reference'] == 'interpass' and main_dict[i+1]['reference'] == 'interpass':
            storage_dict['rho_m'].append(main_dict[i]['rho_m'][-1])
            storage_dict['rho_def'].append(main_dict[i]['rho_def'][-1])
            storage_dict['r_rx'].append(0.5 * main_dict[i]['d_mean'][-1])
            storage_dict['n_rx'].append(main_dict[i]['n_rx'][-1])
            storage_dict['D_def'].append(main_dict[i]['D_def'][-1])
            storage_dict['r_g_curr'].append(main_dict[i]['r_g_curr'][-1])
            storage_dict['n_d'].append((variable_pass['n_d'][-1]))
            'added to make the lengths of the lists equal i.e x_curr:global_time or time = 1:1'
            storage_dict['x_curr'].append(main_dict[i]['x_curr'][-1])
            n_d = variable_pass['n_d'][-1]

        'interpass to deformation'
        if variable_pass['reference'] == 'interpass' and main_dict[i+1]['reference'] == 'deformation':
            storage_dict['n_rx'].append(0.0)
            storage_dict['r_rx'].append(0.0)
            storage_dict['r_g_curr'].append(0.0)
            storage_dict['x_curr'].append(0.0)
            storage_dict['x_c_curr'].append(0.0)
            storage_dict['rho_def'].append(main_dict[i]['rho_m'][-1])
            storage_dict['rho_m'].append(main_dict[i]['rho_m'][-1])
            storage_dict['D_def'].append(0.0)
            n_d = variable_pass['n_d'][-1]


'converting each and every list to a numpy array'
# for value in main_dict.values():
#     for key, dict_value in value.items():
#         if type(dict_value) == list:
#             arr = numpy.array(dict_value)
#             value[key] = arr
#             del dict_value
'plotting'
plot_parameters = ['x_curr', 'd_mean', 'rho_m', 'n_rx', 'stress', 'r_g_curr']
plotting_step(plot_parameters, main_dict)

'exporting values to an excel file'
# workbook.close()
# filename = '/Completed_Schedule.xlsx'
# filepath = os.getcwd() + filename
# workbook.save(filepath)

# name_of_file = 'Complete_Schedule'
# export_to_excel(main_dict, name_of_file)

end = time.time()
print('Time elapsed', end - start)
