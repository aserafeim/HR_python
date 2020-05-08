from deformation import *
import dynamic_recrystallisation as drx
import material_constants as mat_const
from optimization_helper_numpy_implemented import find_target_stress, read_stress_strain_file
import inputdata
from inputdata import fitting_params as fp
import openpyxl
import numpy as np
import time
import os
from decimal import Decimal


location1 = r'D:\HotRolling-env\HR_python-aserafeim-Dict_based_0_def\DP600_single_hit.xlsx'
wb1 = openpyxl.load_workbook(location1)
sheet_curr = wb1['exp.result']

start = time.time()
'making target_strains array outside to access function find_target_stress'
sample = []
for k in range(1, 10):
    sample.append(k * 0.05)

target_strains = numpy.array(sample)

stress_strain_dict = read_stress_strain_file(sheet_curr)
temp_list, strain_rate_list = [], []
for keys in stress_strain_dict.keys():
    temperature = float(keys[:keys.index('-')])
    strain_rate = float(keys[keys.index('-')+1:])
    stress_strain_dict[keys]['target_stress_exp'] = find_target_stress(sheet_curr, target_strains, temperature, strain_rate)
    if temperature not in temp_list:
        temp_list.append(temperature)
    if strain_rate not in strain_rate_list:
        strain_rate_list.append(strain_rate)


def error_fun(x_exp, x_sim):
    # print("simulated: ", x_sim)
    error = 0.0
    for i in range(len(x_exp)):
        error += (x_exp[i] - (x_sim[i]/math.pow(10, 6))) / x_exp[i]
    error_final = error / len(x_exp)
    return error_final


def optimize_second(c1, c2, c3, c4, c5, c6, c7, c8, c9, c10):
    'optimization function'
    'creating a workbook for exporting values'
    ''' r_rx_prev, r_g_prev and the rest of the variables should not be used in this function,
    only the c values will be input '''
    time_step = 0.0001
    workbook = openpyxl.Workbook()

    final_strain = 0.8
    error_arr = []
    for temperature in temp_list:

        for strain_rate in strain_rate_list:
            strain_step = time_step * strain_rate
            length = int(final_strain/strain_step)+1
            stress_list_model = numpy.empty([length])
            strain_list_model = numpy.empty([length])
            stress_target = []
            r_g_prev = 0.0
            n_rx_prev = 0.0
            x_prev = 0.0
            x_c_prev = 0.0
            m = inputdata.M
            strain = 0.0
            i = 0
            'to initialise with a non-zero value for running the simulation'
            rho = 1E+11
            count = 0
            # while round(strain, 5) < final_strain:
            while Decimal(str(strain)) <= Decimal(str(final_strain)):
                count += 1
                deform_ans = calc_deformation_dislocation(rho, time_step, temperature, strain,
                                                       strain_rate, c1, c2, c3, c4,
                                                       c5, c6, c7, c8, c9, c10, m)
                rho_def = deform_ans['rho_curr']
                D_def = deform_ans['d_def']

                'calculation of critical dislocation density'
                flag = False
                n_z = 0.1 if strain_rate > 1 else 1
                part1 = (16 * mat_const.calc_austenite_gb_energy(temperature) *
                         math.pow(strain_rate, n_z) * m) \
                        / (3 * mat_const.calc_burgers_vector(temperature) *
                           mat_const.calc_gb_mobility2_drx(temperature) *
                           math.pow(mat_const.calc_dislocation_line_energy(temperature), 2))

                part2 = (c1 / D_def) + (c2 * math.sqrt(inputdata.rho_const))
                rho_critical = math.pow(part1 * part2, 1 / 3)
                rho_m = rho_def

                if rho_def > rho_critical:
                    temp_variable = drx.dynamic_rx(rho_critical, rho_def, n_rx_prev,
                                                   r_g_prev, x_prev, x_c_prev, time_step, temperature,
                                                   strain_rate, D_def, c1, c2, c10, m)
                    rho_m = temp_variable['rho_m']
                    flag = True
                stress_curr = mat_const.calculate_stress(temperature, strain_rate, rho_m, c7, c8, c9, m)
                stress_list_model[i] = stress_curr
                strain_list_model[i] = strain

                'separate stress on a separate list (stress_target) when target strain is reached'
                for x in target_strains:
                    if round(strain, 3) == round(x, 3):
                        stress_target.append(stress_curr)

                'Updating values at the end of each subroutine'
                if flag:
                    n_rx_prev = temp_variable['n_rx']
                    x_prev = temp_variable['x_curr']
                    x_c_prev = temp_variable['x_c_curr']
                    r_g_prev = temp_variable['r_g_curr']
                rho = rho_m
                i += 1
                strain += strain_step
            "---------------------Fetch experimental data--------------------------------"
            'fetch experimental data for target strains based on temperature and strain rate from dictionary'
            'to do and will give stress_list_exp'
            'error function calculation'
            stress_list_exp = stress_strain_dict[str(temperature)+'-'+str(strain_rate)]['target_stress_exp']

            "printing stress values"
            # print("experimental stresses are: ", stress_list_exp)

            'writing stress and strain values to an excel file'
            worksheet_name = str(temperature)+"-"+str(strain_rate)
            workbook.create_sheet(worksheet_name)
            ws = workbook[worksheet_name]
            ws.cell(row=1, column=1, value="Strain")
            ws.cell(row=1, column=2, value="Stress")
            for i in range(1, len(strain_list_model)):
                ws.cell(row=i+1, column=1, value=strain_list_model[i])
                ws.cell(row=i+1, column=2, value=stress_list_model[i]/math.pow(10, 6))

            workbook.close()
            "calculation of error"
            stress_target_arr = numpy.array(stress_target)
            loc_error = error_fun(stress_list_exp, stress_target_arr)
            error_arr.append(abs(loc_error))
            print("error: ", loc_error, "\n-----------------------------------------------------------------\n")
    filepath = os.getcwd() + "/optimization_result.xlsx"
    # counter_for_name = 0
    # while os.path.exists("optimization_results%s.xlsx" % counter_for_name):
    #     counter_for_name += 1
    # full_filepath = filepath + "optimization_result.xlsx"
    workbook.save(filepath)
    print("error array: ", error_arr)
    error_final = np.sum(error_arr) / np.size(error_arr)
    return error_final


error_return = optimize_second(fp['C_1'], fp['C_2'], fp['C_3'], fp['C_4'], fp['C_5'], fp['C_6'], fp['C_7'], fp['C_8'],
                            fp['C_9'], fp['C_10'])
print(error_return)
end = time.time()
print("time required: ", end-start)
