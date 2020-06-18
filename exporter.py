import openpyxl
import os


def export_to_excel(reference_dict, name):
    workbook = openpyxl.Workbook()
    for pass_key, pass_value in reference_dict.items():
        worksheet_name = "Pass_"+str(pass_key)
        workbook.create_sheet(worksheet_name)
        ws = workbook[worksheet_name]
        j = 1
        for key, value in pass_value.items():

            ws.cell(row=1, column=j, value=key)
            if type(value) == list:
                for i in range(0, len(value)):
                    ws.cell(row=i + 2, column=j, value=value[i])
                    # ws.cell(row=i + 1, column=2, value=list_y[i])
            else:
                ws.cell(row=2, column=j, value= value)
            j += 1
    workbook.close()
    filename = '/'+name+'.xlsx'
    filepath = os.getcwd() + filename
    workbook.save(filepath)


def export_pass(reference_dict, name):
    workbook = openpyxl.Workbook()
    worksheet_name = "Details_" + name
    workbook.create_sheet(worksheet_name)
    ws = workbook[worksheet_name]
    j = 1
    for key, value in reference_dict.items():

        ws.cell(row=1, column=j, value=key)
        if type(value) == list:
            for i in range(0, len(value)):
                ws.cell(row=i + 2, column=j, value=value[i])
        else:
            ws.cell(row=2, column=j, value=value)
        j += 1
    workbook.close()
    filename = '/' + name + '.xlsx'
    filepath = os.getcwd() + filename
    workbook.save(filepath)
