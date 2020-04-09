# -*- coding: utf-8 -*-
"""
Created on Wed Mar 25 21:30:38 2020

@author: aserafeim
"""
import openpyxl

'''define error and experimental array
Need a dictionary with key'''


def read_stress_strain_file(sheet_curr):
    ''' reading the file and storing the stress strain values into the stress-strain dictionary '''
    stress_strain_dict = {}
    '''storing data as stress_strain_dict[temp-strain_rate]= {
                                                stress = [list of values],
                                                strain = [list of values]
                                                }'''
    for i in range(1, 26, 3):
        max_row_for_c = max((c.row for c in sheet_curr[chr(64+i)] if c.value is not None))
        key = str(sheet_curr.cell(1, i).value)+'-'+str(sheet_curr.cell(1, i+1).value)
        sub_dictionary = {'stress': [],
                          'strain': []
        }
        # for row in sheet_curr.iter_rows(min_row=2, max_row=max_row_for_c, values_only=True):
        for j in range(2, max_row_for_c+1):
            sub_dictionary['stress'].append(sheet_curr.cell(j, i+1).value)
            sub_dictionary['strain'].append(sheet_curr.cell(j, i).value)
        stress_strain_dict[key] = sub_dictionary

    return stress_strain_dict


def find_target_stress(target_strains, temperature, strain_rate):
    location1 = r'D:\HotRolling-env\HR_python-aserafeim-Dict_based_0_def\DP600_single_hit.xlsx'
    wb1 = openpyxl.load_workbook(location1)
    sheet_curr1 = wb1['exp.result']
    stress_strain_dict = read_stress_strain_file(sheet_curr1)
    target_stresses = []
    print(target_strains)
    check_key = str(str(temperature)+'-'+str(strain_rate))
    temp_dict = stress_strain_dict[check_key]
    for strain in target_strains:
        if strain in temp_dict['strain']:
            index = temp_dict['strain'].index(strain)
            # print("index: ", index)
            target_stresses.append(temp_dict['stress'][index])
        else:
            a = []
            for i in temp_dict['strain']:
                if round(i, 2) == round(strain, 2):
                    a.append(i)
            # print(a)
            actual_strain = min(a)
            index = temp_dict['strain'].index(actual_strain)
            # print("index: ", index)
            target_stresses.append(temp_dict['stress'][index])
    print("target_stress: ", target_stresses)
    return target_stresses


def error_fun(x_exp, x_sim):
    error = 0
    for i in range(len(x_exp)):
        error += (x_exp[i] - x_sim[i]) / x_exp[i]
    error_final = error/len(x_exp)
    return error_final


def read_parameters_file():
    location = r'D:\HotRolling-env\HR_python-aserafeim-Dict_based_0_def\Test_srx.xlsx'
    wb = openpyxl.load_workbook(location)
    sheet_curr = wb['Static_RX']
    max_row = sheet_curr.max_row
    main_dict = {}
    '''reading excel file'''
    for i in range(3, max_row + 1):
        temp_dict = {}
        temp_dict['time_increment'] = sheet_curr.cell(i, 2).value
        temp_dict['initial_strain'] = sheet_curr.cell(i, 3).value
        temp_dict['final_strain'] = sheet_curr.cell(i, 4).value
        temp_dict['initial_strain_rate'] = sheet_curr.cell(i, 5).value
        temp_dict['final_strain_rate'] = sheet_curr.cell(i, 6).value
        temp_dict['initial_temp'] = sheet_curr.cell(i, 7).value + 273.15
        temp_dict['final_temp'] = sheet_curr.cell(i, 8).value + 273.15
        temp_dict['rho_0'] = sheet_curr.cell(i, 9).value if (isinstance(sheet_curr.cell(i, 9).value, int) or
                                                             isinstance(sheet_curr.cell(i, 9).value, float)) else 0.0
        temp_dict['pass_interval'] = sheet_curr.cell(i, 10).value
        temp_dict['temp_rate_of_change'] = (temp_dict['final_temp'] - temp_dict['initial_temp'])/temp_dict['pass_interval']
        ''
        if temp_dict['initial_strain_rate'] != 0.0:
            temp_dict['reference'] = 'deformation'
        else:
            temp_dict['reference'] = 'interpass'
        main_dict[i - 2] = temp_dict
    # for key, items in main_dict.items():
    #     print(key, ':', items, '\n')

    return main_dict

location1 = r'D:\HotRolling-env\HR_python-aserafeim-Dict_based_0_def\DP600_single_hit.xlsx'
wb1 = openpyxl.load_workbook(location1)
sheet_curr1 = wb1['exp.result']
stress_strain_dict = read_stress_strain_file(sheet_curr1)

