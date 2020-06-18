# -*- coding: utf-8 -*-
"""
Created on Thu Jan 23 13:38:31 2020

@author: sroy
"""
import matplotlib.pyplot as plt
import openpyxl
import os


def plot(x_values, y_values, x_axis_name, y_axis_name, log_x, log_y, reference_dict):
    list_x = []
    list_y = []
    workbook = openpyxl.Workbook()
    worksheet_name = x_axis_name + "-" + y_axis_name
    workbook.create_sheet(worksheet_name)
    ws = workbook[worksheet_name]
    ws.cell(row=1, column=1, value=x_axis_name)
    ws.cell(row=1, column=2, value=y_axis_name)
    for value in reference_dict.values():
        if y_values in value.keys() and x_values in value.keys():
            list_y.extend(value[y_values])
            list_x.extend(value[x_values])
            for i in range(1, len(list_x)):
                ws.cell(row=i + 1, column=1, value=list_x[i])
                ws.cell(row=i + 1, column=2, value=list_y[i])
            workbook.close()
            filename = '/hot_rolling_' + y_axis_name + '-' + x_axis_name + '.xlsx'
            filepath = os.getcwd() + filename
            workbook.save(filepath)
    plt.plot(list_x, list_y)

    if log_x:
        plt.xscale('log')
    if log_y:
        plt.yscale('log')
    plt.grid(True, which="both", linestyle='--')
    plt.xlabel(x_values)
    plt.ylabel(y_values)
    chart_title = y_axis_name+' vs ' + x_axis_name
    plt.title(chart_title)

    plt.show()


def plot_step_wise(y_values, y_axis_name, log_y, reference_dict):
    workbook = openpyxl.Workbook()
    list_x = []
    list_y = []
    # for yvalue in y_values:
    #     worksheet_name = yvalue
    #     workbook.create_sheet(worksheet_name)
    #     ws = workbook[worksheet_name]
    #     ws.cell(row=1, column=1, value=x_axis_name)
    #     ws.cell(row=1, column=2, value=y_axis_name)
    for key, value in reference_dict.items():
        # if y_values in value.keys():
        # list_y.extend(value[y_values])
        # list_x.extend([key]*len(value[y_values]))
        list_x.append(key)
        list_y.append(value[y_values][-1])
    plt.plot(list_x, list_y)
    if log_y:
        plt.yscale('log')
    plt.grid(True, which="both", linestyle='--')
    plt.xlabel("Step")
    plt.ylabel(y_values)
    chart_title = y_axis_name
    plt.title(chart_title)

    plt.show()


def plotting_step(y_values, reference_dict):
    list_x = []
    plotter_helper_y = []
    for keys in reference_dict.keys():
        list_x.append(keys)
    for y_value in y_values:
        list_y = []
        for value in reference_dict.values():
            list_y.append(value[y_value][-1])
        plotter_helper_y.append((list_y, y_value))

    for i in range(len(plotter_helper_y)):
        plt.figure(i+1)
        plt.plot(list_x, plotter_helper_y[i][0])
        plt.grid(True, which="both", linestyle='--')
        plt.xlabel("Step")
        plt.ylabel(plotter_helper_y[i][1])
        plt.title(plotter_helper_y[i][1])
    plt.show()
