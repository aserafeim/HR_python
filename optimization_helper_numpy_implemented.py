# -*- coding: utf-8 -*-
"""
Created on Wed Mar 25 21:30:38 2020

@author: aserafeim
"""
import openpyxl
import numpy
import scipy
from scipy import signal

'''define error and experimental array
Need a dictionary with key'''


def slicing_from_yp(array, yield_point):
    'finds out the index of the yield_point which is approx 0.004 strain'
    # position = 0
    # difference = 1
    # for x in array:
    #     temp_difference = yield_point - x if x < yield_point else x - yield_point
    #     if temp_difference < difference:
    #         difference = temp_difference
    #         position = array.index(x)
    # return position
    index = (numpy.abs(array-yield_point)).argmin()
    return index


# def slice_to_rx_minima(array):
#     # result_peaks = scipy.signal.find_peaks(array)
#     result_peaks = scipy.signal.argrelmin(array)
#     # result_peaks = scipy.signal.
#     return result_peaks


def read_stress_strain_file(sheet_curr):
    ''' reading the file and storing the stress strain values into the stress-strain dictionary '''
    stress_strain_dict = {}
    '''storing data as stress_strain_dict[temp-strain_rate]= {
                                                stress = [list of values],
                                                strain = [list of values]
                                                }'''
    for i in range(1, 26, 3):
        max_row_for_c = max((c.row for c in sheet_curr[chr(64+i)] if c.value is not None))
        key = str(float(sheet_curr.cell(1, i).value)) + '-' + (str(float(sheet_curr.cell(1, i+1).value)))
        sub_dictionary = {'stress': numpy.empty(max_row_for_c),
                          'strain': numpy.empty(max_row_for_c)
                          }

        " choosing only values when stress and strain values are greater than 0"
        for j in range(2, max_row_for_c+1):

            # sub_dictionary['stress'].append(sheet_curr.cell(j, i+1).value)
            # sub_dictionary['strain'].append(sheet_curr.cell(j, i).value)
            sub_dictionary['stress'][j-2] = sheet_curr.cell(j, i+1).value
            sub_dictionary['strain'][j-2] = sheet_curr.cell(j, i).value



        " selecting only the plastic part of the curve, ie. assuming the Y.Point is at a strain of 0.004=4% strain"
        cutoff_yp = slicing_from_yp(sub_dictionary['strain'], 0.004)
        sub_dictionary['strain'] = sub_dictionary['strain'][cutoff_yp:]
        sub_dictionary['stress'] = sub_dictionary['stress'][cutoff_yp:]

        stress_strain_dict[key] = sub_dictionary
    return stress_strain_dict


def find_target_stress(sheet_curr, target_strains, temperature, strain_rate):
    'finds out target stresses from the excel sheet with respect to the target strains, target_stress is a numpy array'
    stress_strain_dict = read_stress_strain_file(sheet_curr)
    temporary = []
    check_key = str(str(temperature) + '-' + str(strain_rate))
    temp_dict = stress_strain_dict[check_key]
    i = 0
    for strain in target_strains:
        if strain in temp_dict['strain']:
            index = numpy.where(temp_dict['strain'] == strain)

        else:
            index = (numpy.abs(temp_dict['strain'] - strain)).argmin()
        temporary.append(temp_dict['stress'][index])

        i += 1
    target_stresses = numpy.array(temporary)
    return target_stresses


def error_fun(x_exp, x_sim):
    error = 0.0
    for i in range(len(x_exp)):
        error += (x_exp[i] - x_sim[i]) / x_exp[i]
    error_final = error/len(x_exp)
    return error_final


def read_parameters_file():
    location = r'D:\HotRolling-env\HR_python-aserafeim-Dict_based_0_def\Test_srx.xlsx'
    wb = openpyxl.load_workbook(location)
    sheet_curr = wb['Static_RX']
    max_row = sheet_curr.max_row
    main_dict = {}
    '''reading excel file'''
    for i in range(3, max_row + 1):
        temp_dict = {}
        temp_dict['time_increment'] = sheet_curr.cell(i, 2).value
        temp_dict['initial_strain'] = sheet_curr.cell(i, 3).value
        temp_dict['final_strain'] = sheet_curr.cell(i, 4).value
        temp_dict['initial_strain_rate'] = sheet_curr.cell(i, 5).value
        temp_dict['final_strain_rate'] = sheet_curr.cell(i, 6).value
        temp_dict['initial_temp'] = sheet_curr.cell(i, 7).value + 273.15
        temp_dict['final_temp'] = sheet_curr.cell(i, 8).value + 273.15
        temp_dict['rho_0'] = sheet_curr.cell(i, 9).value if (isinstance(sheet_curr.cell(i, 9).value, int) or
                                                             isinstance(sheet_curr.cell(i, 9).value, float)) else 0.0
        temp_dict['pass_interval'] = sheet_curr.cell(i, 10).value
        temp_dict['temp_rate_of_change'] = (temp_dict['final_temp'] - temp_dict['initial_temp'])/temp_dict['pass_interval']
        ''
        if temp_dict['initial_strain_rate'] != 0.0:
            temp_dict['reference'] = 'deformation'
        else:
            temp_dict['reference'] = 'interpass'
        main_dict[i - 2] = temp_dict

    return main_dict


# a = numpy.array([0.07, 0.089, 0.003, 0.09, 0.32, 0.43, 0.05, 0.88, 0.10, 0.99, 1.09])
# slice_to_rx_minima(a)
