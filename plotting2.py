import matplotlib.pyplot as plt
import openpyxl
import math
import os


def plot(x_values, y_values, x_axis_name, y_axis_name, log_x, log_y, reference_dict):
    list_x = []
    list_y = []
    workbook = openpyxl.Workbook()
    # worksheet_name = x_axis_name + "-" + y_axis_name
    # workbook.create_sheet(worksheet_name)
    # ws = workbook[worksheet_name]
    # ws.cell(row=1, column=1, value=x_axis_name)
    # ws.cell(row=1, column=2, value=y_axis_name)
    for value in reference_dict.values():
        if y_values in value.keys() and x_values in value.keys():
            list_y.extend(value[y_values])
            list_x.extend(value[x_values])
            # for i in range(1, len(list_x)):
            #     ws.cell(row=i + 1, column=1, value=list_x[i])
            #     ws.cell(row=i + 1, column=2, value=list_y[i])
            # workbook.close()
            # filename = '/hot_rolling_' + y_axis_name + '-' + x_axis_name + '.xlsx'
            # filepath = os.getcwd() + filename
            # workbook.save(filepath)
    plt.plot(list_x, list_y)

    if log_x:
        plt.xscale('log')
    if log_y:
        plt.yscale('log')
    plt.grid(True, which="both", linestyle='--')
    plt.xlabel(x_values)
    plt.ylabel(y_values)
    chart_title = y_axis_name+' vs ' + x_axis_name
    plt.title(chart_title)

    plt.show()

def plot_pass(x_values, y_values, x_axis_name, y_axis_name, log_x, log_y, reference_dict):
    list_x = []
    list_y = []
    # for key, value in reference_dict.items():
    #     if y_values in value.keys() and x_values in value.keys():
    #         list_y.extend(value[y_values])
    #         list_x.extend(value[x_values])
    if x_values in reference_dict and y_values in reference_dict:
        list_y.extend(reference_dict[y_values])
        list_x.extend(reference_dict[x_values])
    plt.plot(list_x, list_y)

    if log_x:
        plt.xscale('log')
    if log_y:
        plt.yscale('log')
    plt.grid(True, which="both", linestyle='--')
    plt.xlabel(x_values)
    plt.ylabel(y_values)
    chart_title = y_axis_name+' vs ' + x_axis_name
    plt.title(chart_title)

    plt.show()